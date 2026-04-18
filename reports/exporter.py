"""Export Excel professionnel avec 6 feuilles et formatage conditionnel."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from loguru import logger
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from config import settings
from data.cache import Signal, OHLCV, SessionLocal
from ml.trainer import load_models


HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F3A5F")
BORDER = Border(
    left=Side(style="thin", color="BBBBBB"),
    right=Side(style="thin", color="BBBBBB"),
    top=Side(style="thin", color="BBBBBB"),
    bottom=Side(style="thin", color="BBBBBB"),
)


def _style_header(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _autosize(ws, df: pd.DataFrame) -> None:
    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(x)) for x in df[col].astype(str).values[:200]] + [10])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 12), 40)


def _write_df(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    _style_header(ws, start_row, len(df.columns))
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws, df)
    return start_row + len(df) + 1


def _signals_to_df(signals: list[Signal]) -> pd.DataFrame:
    return pd.DataFrame([{
        "Ticker": s.ticker,
        "Date": s.created_at,
        "Prix": s.price,
        "Score Consensus": s.consensus_score,
        "Confiance": s.confidence,
        "RF": s.rf_prob,
        "XGBoost": s.xgb_prob,
        "LightGBM": s.lgb_prob or 0.0,
        "LSTM": s.lstm_prob,
        "IsoForest": s.iso_score,
        "Cible +50%": s.target_1,
        "Cible +100%": s.target_2,
        "Cible +200%": s.target_3,
        "Stop-Loss": s.stop_loss,
        "R/R": s.risk_reward,
        "Top Features": s.top_features,
    } for s in signals])


def _build_dashboard_sheet(ws, signals: list[Signal], metrics: dict) -> None:
    ws["A1"] = "INVESTMENT SNIPER — Tableau de bord"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")

    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    ws["A2"] = f"Généré : {now}"
    ws["A2"].font = Font(italic=True, color="555555")

    ws["A4"] = "KPI"; ws["B4"] = "Valeur"
    kpis = [
        ("Signaux totaux", len(signals)),
        ("ULTRA", sum(1 for s in signals if s.confidence == "ULTRA")),
        ("HIGH", sum(1 for s in signals if s.confidence == "HIGH")),
        ("MEDIUM", sum(1 for s in signals if s.confidence == "MEDIUM")),
        ("Score moyen", round(sum(s.consensus_score for s in signals) / len(signals), 4) if signals else 0),
        ("Seuil consensus", settings.consensus_threshold),
        ("Tickers suivis", metrics.get("dataset", {}).get("n_tickers", 0)),
        ("Échantillons entraînement", metrics.get("dataset", {}).get("n_rows", 0)),
    ]
    for i, (k, v) in enumerate(kpis, start=5):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    _style_header(ws, 4, 2)

    ws["D4"] = "Modèle"; ws["E4"] = "Accuracy"; ws["F4"] = "Precision"; ws["G4"] = "Recall"; ws["H4"] = "F1"; ws["I4"] = "ROC-AUC"
    row = 5
    for key in ["random_forest", "xgboost", "lstm"]:
        m = metrics.get(key, {})
        ws.cell(row=row, column=4, value=m.get("name", key))
        ws.cell(row=row, column=5, value=round(m.get("accuracy", 0), 4))
        ws.cell(row=row, column=6, value=round(m.get("precision", 0), 4))
        ws.cell(row=row, column=7, value=round(m.get("recall", 0), 4))
        ws.cell(row=row, column=8, value=round(m.get("f1", 0), 4))
        ws.cell(row=row, column=9, value=round(m.get("roc_auc", 0), 4))
        row += 1
    _style_header(ws, 4, 9)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    for col in "DEFGHI":
        ws.column_dimensions[col].width = 14


def _build_active_signals_sheet(ws, signals: list[Signal]) -> None:
    df = _signals_to_df(signals)
    if df.empty:
        ws["A1"] = "Aucun signal actif"
        return
    end = _write_df(ws, df)

    score_col = get_column_letter(df.columns.get_loc("Score Consensus") + 1)
    ws.conditional_formatting.add(
        f"{score_col}2:{score_col}{end}",
        ColorScaleRule(
            start_type="num", start_value=0.5, start_color="F8D7DA",
            mid_type="num", mid_value=0.75, mid_color="FFF3CD",
            end_type="num", end_value=1.0, end_color="D4EDDA",
        ),
    )
    rr_col = get_column_letter(df.columns.get_loc("R/R") + 1)
    ws.conditional_formatting.add(
        f"{rr_col}2:{rr_col}{end}",
        CellIsRule(operator="greaterThanOrEqual", formula=["2"],
                   fill=PatternFill("solid", fgColor="C6EFCE")),
    )


def _build_history_sheet(ws, signals: list[Signal]) -> None:
    df = _signals_to_df(signals)
    if df.empty:
        ws["A1"] = "Historique vide"
        return
    _write_df(ws, df)


def _build_ml_sheet(ws, metrics: dict) -> None:
    ws["A1"] = "Analyse ML"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    row = 3
    ws.cell(row=row, column=1, value="Modèle")
    ws.cell(row=row, column=2, value="Accuracy")
    ws.cell(row=row, column=3, value="Precision")
    ws.cell(row=row, column=4, value="Recall")
    ws.cell(row=row, column=5, value="F1")
    ws.cell(row=row, column=6, value="ROC-AUC")
    _style_header(ws, row, 6)
    row += 1
    for key in ["random_forest", "xgboost", "lstm"]:
        m = metrics.get(key, {})
        ws.cell(row=row, column=1, value=m.get("name", key))
        ws.cell(row=row, column=2, value=round(m.get("accuracy", 0), 4))
        ws.cell(row=row, column=3, value=round(m.get("precision", 0), 4))
        ws.cell(row=row, column=4, value=round(m.get("recall", 0), 4))
        ws.cell(row=row, column=5, value=round(m.get("f1", 0), 4))
        ws.cell(row=row, column=6, value=round(m.get("roc_auc", 0), 4))
        row += 1

    chart = BarChart()
    chart.title = "Comparaison des modèles"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Modèle"
    data = Reference(ws, min_col=2, min_row=3, max_col=6, max_row=row - 1)
    cats = Reference(ws, min_col=1, min_row=4, max_row=row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10; chart.width = 20
    ws.add_chart(chart, "H3")

    row += 3
    ws.cell(row=row, column=1, value="Feature Importance (Random Forest)").font = Font(bold=True, size=13)
    row += 1
    ws.cell(row=row, column=1, value="Feature"); ws.cell(row=row, column=2, value="Importance")
    _style_header(ws, row, 2)
    row += 1
    imp_rf = metrics.get("feature_importance", {}).get("random_forest", {})
    for name, val in sorted(imp_rf.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=round(val, 5))
        row += 1


def _build_raw_data_sheet(ws, limit_per_ticker: int = 200) -> None:
    ws["A1"] = "Données OHLCV (cache)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")
    with SessionLocal() as session:
        rows = session.query(OHLCV).order_by(OHLCV.ticker, OHLCV.date.desc()).limit(5000).all()
    if not rows:
        ws["A3"] = "Pas de données cachées"
        return
    df = pd.DataFrame([{
        "Ticker": r.ticker, "Date": r.date,
        "Open": r.open, "High": r.high, "Low": r.low, "Close": r.close,
        "Adj Close": r.adj_close, "Volume": r.volume,
    } for r in rows])
    _write_df(ws, df, start_row=3)


def _build_params_sheet(ws) -> None:
    ws["A1"] = "Paramètres système"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    params = [
        ("App env", settings.app_env),
        ("Log level", settings.log_level),
        ("DB path", str(settings.db_path)),
        ("Models dir", str(settings.models_dir)),
        ("Scan interval (min)", settings.scan_interval_minutes),
        ("Consensus threshold", settings.consensus_threshold),
        ("Lookback days", settings.lookback_days),
        ("History years", settings.history_years),
        ("Min gain %", settings.min_gain_pct),
        ("SMTP host", settings.smtp_host or "—"),
        ("SMTP user", settings.smtp_user or "—"),
        ("Alert email", settings.alert_to_email or "—"),
    ]
    ws["A3"] = "Paramètre"; ws["B3"] = "Valeur"
    _style_header(ws, 3, 2)
    for i, (k, v) in enumerate(params, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42


def export_report(output_path: Path | None = None) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    with SessionLocal() as session:
        all_signals = session.query(Signal).order_by(Signal.created_at.desc()).all()

    active = [s for s in all_signals if s.consensus_score >= settings.consensus_threshold]
    metrics = load_models().get("metrics", {})

    _build_dashboard_sheet(ws, active, metrics)
    _build_active_signals_sheet(wb.create_sheet("Signaux Actifs"), active)
    _build_history_sheet(wb.create_sheet("Historique"), all_signals)
    _build_ml_sheet(wb.create_sheet("Analyse ML"), metrics)
    _build_raw_data_sheet(wb.create_sheet("Données Brutes"))
    _build_params_sheet(wb.create_sheet("Paramètres"))

    output_path = output_path or (
        settings.exports_dir / f"sniper_report_{datetime.utcnow():%Y%m%d_%H%M%S}.xlsx"
    )
    wb.save(output_path)
    logger.info(f"[export] Rapport sauvegardé : {output_path}")
    return output_path
